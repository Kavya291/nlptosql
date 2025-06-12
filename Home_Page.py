import streamlit as st
import sqlite3
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Upload Excel", layout="centered")
st.markdown("<style>[data-testid='stSidebar'] { display: none; }</style>", unsafe_allow_html=True)

st.title("📤 Upload Excel to Populate Student DB")

uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type="xlsx")

if "upload_success" not in st.session_state:
    st.session_state.upload_success = False

def read_excel(file):
    # Load workbook
    wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
    sheet = wb.active

    # Read header row
    headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in next(sheet.iter_rows(max_row=1))]

    # Read all data rows as list of dicts
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    return headers, data

if uploaded_file:
    try:
        headers, data = read_excel(uploaded_file)

        required_columns = ['Name', 'CGPA', 'Location', 'Email', 'Phone Number', 'Preferred Work Location', 'Specialization in Degree']
        missing_cols = [col for col in required_columns if col not in headers]

        if missing_cols:
            st.error(f"❌ Missing required columns: {', '.join(missing_cols)}")
        else:
            conn = sqlite3.connect("students.db")
            cursor = conn.cursor()

            # Create table if not exists
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT,
                    cgpa REAL,
                    location TEXT,
                    email TEXT,
                    phone_number TEXT,
                    preferred_work_location TEXT,
                    specialization TEXT
                )
            ''')

            # Delete existing records before inserting new ones
            cursor.execute('DELETE FROM students')

            inserted = 0
            for idx, row in enumerate(data):
                try:
                    cursor.execute('''
                        INSERT INTO students (
                            name, cgpa, location, email, phone_number, preferred_work_location, specialization
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        row.get('Name'),
                        float(row.get('CGPA')) if row.get('CGPA') is not None else None,
                        row.get('Location'),
                        row.get('Email'),
                        str(row.get('Phone Number')),
                        row.get('Preferred Work Location'),
                        row.get('Specialization in Degree')
                    ))
                    inserted += 1
                except Exception as e:
                    st.warning(f"⚠️ Skipped row {idx + 2}: {e}")

            conn.commit()
            conn.close()

            st.success(f"✅ Successfully inserted {inserted} records.")
            st.session_state.upload_success = True

    except Exception as e:
        st.error(f"❌ Failed to process the uploaded file: {e}")
else:
    st.info("⬆️ Please upload an Excel file.")

if st.session_state.upload_success:
    st.markdown("---")
    st.success("🎉 Upload complete! You can now query the data.")
    if st.button("Go to Query Page"):
        st.switch_page("pages/2_Query_Database.py")

